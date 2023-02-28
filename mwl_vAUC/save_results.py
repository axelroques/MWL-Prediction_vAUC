import xlsxwriter
import os
from openpyxl import workbook
from openpyxl import load_workbook

def write_in_table(X, ground_truth, name, auc_mean, auc_std):
    
    initial_path = "C:\\Users\\Utilisateur\\Desktop\\MWL_project_Alice\\mwl-project\\results"
    directory_path = os.path.join(initial_path, ground_truth)
    
    if not os.path.exists(directory_path):
        os.makedirs(directory_path)
    
#    file_path = os.path.join(directory_path, name + '.xlsx')
    file_path = os.path.join(directory_path, "results" + '.xlsx')
    
    if not os.path.exists(file_path):
        file = xlsxwriter.Workbook(file_path)
        file.close()
    file = load_workbook(file_path)
    sheets = file.sheetnames
    Sheet1 = file[sheets[0]]
    
    Sheet1.cell(row = 1, column = 1).value = ground_truth
    Sheet1.cell(row = 1, column = 3).value = "Pilot"
    Sheet1.cell(row = 2, column = 1).value = "Features used"
    Sheet1.cell(row = 2, column = 2).value = "Mean"
    
    for k in range(7):
        Sheet1.cell(row = 2, column = k+3).value = k+3
    
    if name == "whole dataset":
        Sheet1.cell(row = 3, column = 1).value = name
        Sheet1.cell(row = 3, column = 2).value = "%.3f"%auc_mean + " ±" + "%.3f"%auc_std
        for k in range (3, 10):
            Sheet1.cell(row = 3, column = k).value = "%.3f"%X[k-3,0] + " [" + "%.3f"%X[k-3,1] + ", " + "%.3f"%X[k-3,2] + "]"  
        file.save(file_path)
        
    if name == "helico data removed":
        Sheet1.cell(row = 4, column = 1).value = name
        Sheet1.cell(row = 4, column = 2).value = "%.3f"%auc_mean + " ±" + "%.3f"%auc_std
        for k in range (3, 10):
            Sheet1.cell(row = 4, column = k).value = "%.3f"%X[k-3,0] + " [" + "%.3f"%X[k-3,1] + ", " + "%.3f"%X[k-3,2] + "]"  
        file.save(file_path)
        
    if name == "only helico data":
        Sheet1.cell(row = 5, column = 1).value = name
        Sheet1.cell(row = 5, column = 2).value = "%.3f"%auc_mean + " ±" + "%.3f"%auc_std
        for k in range (3, 10):
            Sheet1.cell(row = 5, column = k).value = "%.3f"%X[k-3,0] + " [" + "%.3f"%X[k-3,1] + ", " + "%.3f"%X[k-3,2] + "]"  
        file.save(file_path)
        
    if name == "FC removed":
        Sheet1.cell(row = 6, column = 1).value = name
        Sheet1.cell(row = 6, column = 2).value = "%.3f"%auc_mean + " ±" + "%.3f"%auc_std
        for k in range (3, 10):
            Sheet1.cell(row = 6, column = k).value = "%.3f"%X[k-3,0] + " [" + "%.3f"%X[k-3,1] + ", " + "%.3f"%X[k-3,2] + "]"  
        file.save(file_path)
        
    if name == "only flight commands":
        Sheet1.cell(row = 7, column = 1).value = name
        Sheet1.cell(row = 7, column = 2).value = "%.3f"%auc_mean + " ±" + "%.3f"%auc_std
        for k in range (3, 10):
            Sheet1.cell(row = 7, column = k).value = "%.3f"%X[k-3,0] + " [" + "%.3f"%X[k-3,1] + ", " + "%.3f"%X[k-3,2] + "]"  
        file.save(file_path)
        
    if name == "helico+FC removed":
        Sheet1.cell(row = 8, column = 1).value = name
        Sheet1.cell(row = 8, column = 2).value = "%.3f"%auc_mean + " ±" + "%.3f"%auc_std
        for k in range (3, 10):
            Sheet1.cell(row = 8, column = k).value = "%.3f"%X[k-3,0] + " [" + "%.3f"%X[k-3,1] + ", " + "%.3f"%X[k-3,2] + "]"  
        file.save(file_path)
    
    if name == "only helico+FC":
        Sheet1.cell(row = 9, column = 1).value = name
        Sheet1.cell(row = 9, column = 2).value = "%.3f"%auc_mean + " ±" + "%.3f"%auc_std
        for k in range (3, 10):
            Sheet1.cell(row = 9, column = k).value = "%.3f"%X[k-3,0] + " [" + "%.3f"%X[k-3,1] + ", " + "%.3f"%X[k-3,2] + "]"  
        file.save(file_path)
        
    if name == "oculo removed":
        Sheet1.cell(row = 10, column = 1).value = name
        Sheet1.cell(row = 10, column = 2).value = "%.3f"%auc_mean + " ±" + "%.3f"%auc_std
        for k in range (3, 10):
            Sheet1.cell(row = 10, column = k).value = "%.3f"%X[k-3,0] + " [" + "%.3f"%X[k-3,1] + ", " + "%.3f"%X[k-3,2] + "]"  
        file.save(file_path)
        
    if name == "only oculo":
        Sheet1.cell(row = 11, column = 1).value = name
        Sheet1.cell(row = 11, column = 2).value = "%.3f"%auc_mean + " ±" + "%.3f"%auc_std
        for k in range (3, 10):
            Sheet1.cell(row = 11, column = k).value = "%.3f"%X[k-3,0] + " [" + "%.3f"%X[k-3,1] + ", " + "%.3f"%X[k-3,2] + "]"  
        file.save(file_path)
        
    if name == "oculo+AOI removed":
        Sheet1.cell(row = 12, column = 1).value = name
        Sheet1.cell(row = 12, column = 2).value = "%.3f"%auc_mean + " ±" + "%.3f"%auc_std
        for k in range (3, 10):
            Sheet1.cell(row = 12, column = k).value = "%.3f"%X[k-3,0] + " [" + "%.3f"%X[k-3,1] + ", " + "%.3f"%X[k-3,2] + "]"  
        file.save(file_path)
        
    if name == "only oculo+AOI":
        Sheet1.cell(row = 13, column = 1).value = name
        Sheet1.cell(row = 13, column = 2).value = "%.3f"%auc_mean + " ±" + "%.3f"%auc_std
        for k in range (3, 10):
            Sheet1.cell(row = 13, column = k).value = "%.3f"%X[k-3,0] + " [" + "%.3f"%X[k-3,1] + ", " + "%.3f"%X[k-3,2] + "]"  
        file.save(file_path)
        
    if name == "AOI removed":
        Sheet1.cell(row = 14, column = 1).value = name
        Sheet1.cell(row = 14, column = 2).value = "%.3f"%auc_mean + " ±" + "%.3f"%auc_std
        for k in range (3, 10):
            Sheet1.cell(row = 14, column = k).value = "%.3f"%X[k-3,0] + " [" + "%.3f"%X[k-3,1] + ", " + "%.3f"%X[k-3,2] + "]"  
        file.save(file_path)
        
    if name == "only AOI":
        Sheet1.cell(row = 15, column = 1).value = name
        Sheet1.cell(row = 15, column = 2).value = "%.3f"%auc_mean + " ±" + "%.3f"%auc_std
        for k in range (3, 10):
            Sheet1.cell(row = 15, column = k).value = "%.3f"%X[k-3,0] + " [" + "%.3f"%X[k-3,1] + ", " + "%.3f"%X[k-3,2] + "]"  
        file.save(file_path)
        
    if name == "neurovégé removed":
        Sheet1.cell(row = 16, column = 1).value = name
        Sheet1.cell(row = 16, column = 2).value = "%.3f"%auc_mean + " ±" + "%.3f"%auc_std
        for k in range (3, 10):
            Sheet1.cell(row = 16, column = k).value = "%.3f"%X[k-3,0] + " [" + "%.3f"%X[k-3,1] + ", " + "%.3f"%X[k-3,2] + "]"  
        file.save(file_path)
        
    if name == "only neurovégé":
        Sheet1.cell(row = 17, column = 1).value = name
        Sheet1.cell(row = 17, column = 2).value = "%.3f"%auc_mean + " ±" + "%.3f"%auc_std
        for k in range (3, 10):
            Sheet1.cell(row = 17, column = k).value = "%.3f"%X[k-3,0] + " [" + "%.3f"%X[k-3,1] + ", " + "%.3f"%X[k-3,2] + "]"  
        file.save(file_path)
        
    if name == "helico+interface+AOI":
        Sheet1.cell(row = 18, column = 1).value = name
        Sheet1.cell(row = 18, column = 2).value = "%.3f"%auc_mean + " ±" + "%.3f"%auc_std
        for k in range (3, 10):
            Sheet1.cell(row = 18, column = k).value = "%.3f"%X[k-3,0] + " [" + "%.3f"%X[k-3,1] + ", " + "%.3f"%X[k-3,2] + "]"  
        file.save(file_path)
        
    if name == "only neurovégé+oculo":
        Sheet1.cell(row = 19, column = 1).value = name
        Sheet1.cell(row = 19, column = 2).value = "%.3f"%auc_mean + " ±" + "%.3f"%auc_std
        for k in range (3, 10):
            Sheet1.cell(row = 19, column = k).value = "%.3f"%X[k-3,0] + " [" + "%.3f"%X[k-3,1] + ", " + "%.3f"%X[k-3,2] + "]"  
        file.save(file_path)